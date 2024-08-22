### Calculate Avg and Std for Each Algorithm 
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import re

# Convert Origin File to Avg and Std
def convertToAS(file_path, save_path, algorithm_num = 11, num_columns = 30):
    file = pd.read_excel(file_path, sheet_name="overall")

    ###
    rows_per_column = algorithm_num

    split_avg_columns = {}
    split_std_columns = {}

    for i in range(num_columns):
        start = i * rows_per_column
        end = (i + 1) * rows_per_column
        column_name = f'Column_{i + 1}'
        split_avg_columns[column_name] = file['mean'].iloc[start:end].reset_index(drop=True)
        split_std_columns[column_name] = file['std'].iloc[start:end].reset_index(drop=True)

    # Concatenate the split columns into a new Dataframe
    avg_col = pd.concat(split_avg_columns.values(), axis=1)
    std_col = pd.concat(split_std_columns.values(), axis=1)

    # To make the 11 rows * 30 cols to 30 rows * 11 cols
    avg_row = avg_col.transpose()
    std_row = std_col.transpose()

    # We should use pd to change the label to the desired index.
    avg_row['index'] = range(0, num_columns * 2, 2)
    std_row['index'] = range(1, num_columns * 2, 2)

    # Concate the avg and std row
    table = pd.concat([avg_row, std_row], axis=0)

    # Sort by the created index to get the table
    table = table.sort_values(by='index')

    table = table.drop('index', axis=1)
    table = table.T
    print(save_path)
    table.to_excel(save_path)

# Divide Tables
def split_groups(file_path, save_path):
    # Load the Excel file
    df = pd.read_excel(file_path)

    # Calculate the number of groups
    num_groups = len(df.columns[1:]) // 6

    # Initialize an empty DataFrame to hold all the groups
    all_groups = pd.DataFrame()

    for i in range(num_groups):
        # Extract the group data
        group_data = df.iloc[:, [0] + list(range(i*6+1, (i+1)*6+1))]

        # Remove the columns that only contain NaN values
        group_data = group_data.dropna(axis=1, how='all')

        # Append the group data to the all_groups DataFrame
        all_groups = pd.concat([all_groups, group_data], axis=1)

    # Reset the index of all_groups DataFrame
    all_groups.reset_index(drop=True, inplace=True)

    # Initialize writer
    writer = pd.ExcelWriter(save_path)

    # Calculate the number of columns in each group
    group_cols = all_groups.shape[1] // num_groups

    for i in range(num_groups):
        # Extract the group data
        group_data = all_groups.iloc[:, i*group_cols:(i+1)*group_cols]
        
        # Create the headers
        begin_index = i * 3 + 1
        headers = pd.DataFrame([['F' + str(begin_index + j // 2) if j % 2 == 1 else '' for j in range(group_cols)], 
                                [''] + ['AVG' if j % 2 != 0 else 'STD' for j in range(1, group_cols)]], 
                                columns=group_data.columns)

        # Add the headers to the group data
        group_data = pd.concat([headers, group_data])

        # Write the group data to the Excel file in the same sheet but different blocks of rows
        group_data.to_excel(writer, sheet_name='Grouped Data', startrow=i*(group_data.shape[0]) + 1, index=False, header=False)

    # Save the Excel file
    print(save_path)
    writer._save()

# Select min and highlight
def highlight_min_values(file_path, algorithms, group_num = 10, row_len = 13):
    # Load the Excel file using openpyxl to modify styles
    wb = load_workbook(file_path)
    ws = wb.active

    algorithm_size = len(algorithms)
    # Calculate max number of rows and columns manually
    max_row = ws.max_row
    max_col = len(list(ws.columns))  # Subtract 1 to exclude the index column if it was saved

    # Iterate over each group and column to find and bold the minimum value
    for group_num in range(group_num - 1):  # There are 10 groups
        for index in range(1, algorithm_size + 1):
            row_index = group_num * row_len + index + 3
            val = ws.cell(row_index, 1).value
            ws.cell(row_index, 1, value=algorithms[val])
            # print(ws.cell(row_index, 1).value)
        for col_idx in range(2, max_col + 1):  # Start from column 2 to skip the first column with indices
            min_row = group_num * row_len + 2  # Skip headers
            min_val = float('inf')
            min_row_val = None

            # Find the minimum value in the current column of the group
            for row in range(min_row, min(min_row + row_len, max_row)):
                cell_val = ws.cell(row=row, column=col_idx).value
                if isinstance(cell_val, (int, float)) and cell_val < min_val:
                    min_val = cell_val
                    min_row_val = row

            # Bold the cell with the minimum value if found within the data range
            if min_row_val is not None:
                ws.cell(row=min_row_val, column=col_idx).font = Font(bold=True)

    print(file_path)
    # Save the changes
    wb.save(file_path)

# Init filepath
file_path = r"D:\Project\MDM\Python\Table\Excel\GDAO.xlsx"
save_path = r"D:\Project\MDM\Python\Table\Excel\\"
file_name = re.search(r'[^\\]+$', file_path).group()
save_name = "Std-" + file_name
save_path = save_path + save_name

# Basic
# algorithms = ['GDAO', 'BA', 'DE', 'FA', 'GWO', 'HHO', 'INFO', 'MFO', 'PSO', 'SCA', 'WOA']
# Improvement
algorithms = ['GDAO', 'OBSCA', 'CAGWO', 'CDLOBA', 'LSHADE', 'IGWO', 'MSCA', 'JADE', 'RDWOA', 'RCBA', 'EBOwithCMAR']
# Ablation
# algorithms = ['GDAO', 'GAO', 'DAO', 'AO']

# The number of algorithms
algorithm_len = len(algorithms)
# The number of benchmark function
function_len = 30
# The two additional rows mean function name and avg-std
row_len = algorithm_len + 2

convertToAS(file_path, save_path, algorithm_len, function_len)
split_groups(save_path, save_path)
highlight_min_values(save_path, algorithms, algorithm_len, row_len)