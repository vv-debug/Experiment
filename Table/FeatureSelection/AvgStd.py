### Calculate Avg and Std for Each Experiment 
import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

### Calculate Avg and Std of Feature Selection Results

# Define the directory
root_dir = r'D:\Project\MDM\Paper\GDAO\Experiment\FeatureSelection\Other'

# Define the file
columns_of_interest = ['Algorithms', 'Best-Fitness', 'Error', 'NumOfFeature']

# Search patterns, used to match Excel files ending with Summary
search_patterns = ['*Summary2.csv', '*Summary.csv']
metrics = ['std', 'mean']
results = [[[], []] for _ in range(3)]


# Function: Extract required data from an Excel file
def extract_data_from_excel(file_path, metric_index):
    # Load the CSV file into a DataFrame
    df = pd.read_csv(file_path)

    # Filter the rows where the 'Metric' column is 'std'
    df_filtered = df[df['Metric'] == metrics[metric_index]]

    # Replace 'B' with 'b' in the 'Algorithms' column
    processed_data = df_filtered[['Algorithms', 'Benchmark', 'Best fitness']].replace('B', 'b', regex=True).values.tolist()
    results[0][metric_index].append(processed_data)

    error_data = df_filtered[['Algorithms', 'Benchmark', 'Error']].values.tolist()
    results[1][metric_index].append(error_data)

    numOfFeatures_data = df_filtered[['Algorithms', 'Benchmark', 'NumOfFeatures']].values.tolist()
    results[2][metric_index].append(numOfFeatures_data)

def get_rank(file_path):
    # Read specific rows from the CSV file into a DataFrame
    df = pd.read_csv(file_path, header=None, encoding='latin-1')

    # Select rows 3, 7, 11 and start from the third column (indexing starts from 0)
    rows = df.iloc[[3, 7, 11], 2:]

    # Convert the rankings to integers
    rows = rows.astype(int)

    # Rank the elements in each row
    rows_ranked = rows.rank(axis=1)

    return rows_ranked

def recal_rank(root_dir):
    sum_ranks = None
    num_files = 0

    for dir_path, dir_names, file_names in os.walk(root_dir):
        for dir_name in dir_names:
            dir_path = os.path.join(root_dir, dir_name, '*Order.csv')
            file_paths = glob.glob(dir_path, recursive=True)
            for file_path in file_paths:
                rank = get_rank(file_path)
                
                # Sum the ranks
                if sum_ranks is None:
                    sum_ranks = rank
                else:
                    sum_ranks += rank
                num_files += 1

    # Calculate the mean ranks
    mean_ranks = (sum_ranks / num_files).round(2)

    # Recalculate the ranking of each element in each row
    final_ranks = mean_ranks.rank(axis=1)
    return mean_ranks, final_ranks

def save_rank(root_dir, file_names):
    mean_ranks, final_ranks = recal_rank(root_dir)
    # Assuming df1, df2, df3 are your dataframes
    df1 = pd.DataFrame()
    df2 = pd.DataFrame()
    df3 = pd.DataFrame()

    # List of dataframes
    dfs = [df1, df2, df3]

    mean_ranks_list = mean_ranks.values.tolist()
    final_ranks_list = final_ranks.values.tolist()

    for i in range(len(dfs)):
        ser1 = pd.Series(mean_ranks_list[i])
        ser2 = pd.Series(final_ranks_list[i])
        dfs[i] = pd.concat([ser1, ser2], axis=1).T
        
        # Load the existing data from the Excel file
        original_df = pd.read_excel(file_names[i], sheet_name='Sheet1')

        # Append the new data to the end of the existing data
        for df in dfs:
            original_df = original_df._append(df, ignore_index=True)

        # Write the combined data back to the Excel file
        original_df.to_excel(file_names[i], sheet_name='Sheet1', index=False)
        print(file_names[i])

for dir_path, dir_names, file_names in os.walk(root_dir):
    for dir_name in dir_names:
        for i in range(len(search_patterns)):
            dir_path = os.path.join(root_dir, dir_name, search_patterns[i])
            file_path = glob.glob(dir_path, recursive=True)[0]
            extract_data_from_excel(file_path, i)

file_names = []
for i in range(len(results)):
    Std = results[i][0]
    Avg = results[i][1]

    # Flatten the lists and create dataframes
    df_std = pd.DataFrame([item for sublist in Std for item in sublist], columns=['Algorithm', 'Dateset', 'Data'])
    df_avg = pd.DataFrame([item for sublist in Avg for item in sublist], columns=['Algorithm', 'Dateset', 'Data'])

    # Add a new column to differentiate between Std and Avg
    df_std['Metric'] = 'Std'
    df_avg['Metric'] = 'Avg'

    # Concatenate the Std and Avg dataframes
    df = pd.concat([df_std, df_avg])

    # Pivot the dataframe
    df_pivot = df.pivot_table(index=['Dateset', 'Metric'], columns='Algorithm', values='Data')

    original_order = df['Algorithm'].unique()
    df_pivot = df_pivot[original_order]

    # Save to Excel
    excel_file = f'Feature-{columns_of_interest[i + 1]}.xlsx'
    file_names.append(excel_file)
    df_pivot.to_excel(excel_file)

    # Load the workbook and select the sheet
    book = load_workbook(excel_file)
    sheet = book.active

    # Apply the font to the cells with the minimum values in each row
    for row in sheet.iter_rows(min_row=2, min_col=3, max_row=sheet.max_row):
        min_cell = min((cell for cell in row if cell.value is not None), key=lambda cell: cell.value)
        min_cell.font = Font(bold=True)
        
    # Save the workbook
    book.save(excel_file)

save_rank(root_dir, file_names)