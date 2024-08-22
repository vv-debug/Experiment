### Highlight All Minimum Value
from openpyxl import load_workbook
from openpyxl.styles import Font

### Bold The Minimum Value In Each Line
def select_data(file_path):
    # Load the workbook
    wb = load_workbook(file_path)

    # Select the worksheet
    ws = wb.active

    # Iterate through each row starting from the first row and each column starting from the third column
    for i, row in enumerate(ws.iter_rows(min_row=2, min_col=3, max_row=25, values_only=True), start=2):
        # Convert the row to a list and remove None values
        row = [value for value in row if value is not None]

        # If the row is not empty
        if row:
            # Find the minimum value
            min_value = min(row)

            # Iterate over each value in the row
            for j, value in enumerate(row, start=3):
                # If the value is equal to the minimum value
                if value == min_value:
                    # Bold the cell
                    cell = ws.cell(row=i, column=j)
                    cell.font = Font(bold=True)

    # Save the workbook
    wb.save(file_path)

file_paths = [r"D:\Project\MDM\Python\Feature-Best-Fitness.xlsx",
              r"D:\Project\MDM\Python\Feature-Error.xlsx", 
              r"D:\Project\MDM\Python\Feature-NumOfFeature.xlsx"]
for file_path in file_paths:
    select_data(file_path)