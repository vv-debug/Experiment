### Combine the Avg and Std Result of Mult-Dim
import numpy as np
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font

def select_data(file_paths, save_path, col_len, function_len):
    file_path = file_paths[0]
    # file_path = "E:\MDM\Python\BenchmarkResultManipulate_v3\Function23-06-13_09_22.xlsx"
    file_name = re.search(r'[^\\]+$', file_path).group()            # Get name
    save_name = "Dim-" + file_name
    save_path = save_path + save_name
    print(save_path)
    
    results = [[] for _ in range(function_len + 2)]
    min_values = [[] for _ in range(function_len + 2)]
    for path_index in range(len(file_paths)):
            
        # Load the Sheet
        sheet_data = pd.read_excel(file_paths[path_index], sheet_name="overall", header=None)
        sheet_size = len(sheet_data)
        for i in range(sheet_size // col_len):
            row_index = i * col_len + 1
            row_avg = sheet_data.iloc[row_index, 4]
            row_std = sheet_data.iloc[row_index, 5]
            if path_index == 0:
                results[0].append(f'F{i + 1}')
                results[0].append('')
                results[1].append('AVG')
                results[1].append('STD')
            results[path_index + 2].append(row_avg)
            results[path_index + 2].append(row_std)

    df = pd.DataFrame(results)
    # Transpose the DataFrame
    df = df.T
    # Write the DataFrame to an Excel file
    df.to_excel(save_path, index=False, header=False)

    # Load the workbook
    wb = load_workbook(save_path)

    # Select the worksheet
    ws = wb.active


    # Iterate through each row starting from the first row and each column starting from the third column
    for i, row in enumerate(ws.iter_rows(min_row=1, min_col=3, values_only=True), start=1):
        # Convert the row to a list and remove None values
        row = [value for value in row if value is not None]

        # If the row is not empty
        if row:
            # Find the minimum value and its column index
            min_value = min(row)
            for j, value in enumerate(row, start=3):
                if value == min_value:
                    # Bold the cell with the minimum value
                    cell = ws.cell(row=i, column=j)
                    cell.font = Font(bold=True)

    # Save the workbook
    wb.save(save_path)

file_paths = [r"D:\Project\MDM\Python\Table\Excel\LGDAO-10.xlsx", 
              r"D:\Project\MDM\Python\Table\Excel\LGDAO-30.xlsx", 
              r"D:\Project\MDM\Python\Table\Excel\LGDAO-50.xlsx", 
              r"D:\Project\MDM\Python\Table\Excel\LGDAO-100.xlsx"]
save_path = r"D:\Project\MDM\Python\Table\Excel\\"

# The number of benchmark function
function_len = 30
# The number of algorithms + 1
col_len = 11
select_data(file_paths, save_path, col_len=col_len, function_len=function_len)
